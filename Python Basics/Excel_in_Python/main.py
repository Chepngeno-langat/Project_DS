import openpyxl

wb = openpyxl.load_workbook('videogamesales.xlsx')

# call an active worksheet
ws = wb['vgsales']

# count no of rows and columns
# print(f"Total number of rows: {ws.max_row}")
# print(f"Total number of columns: {ws.max_column}")

# Reading data from a cell
# print(f"Value in cell B5 is: {ws['B5'].value}")

# Reading Data from multiple cells
# print all cell values in a row
for i in range(1, ws.max_column):
    values = ws.cell(row=1, column=i).value
    # print(values)

# Print multiple rows in a specific column
for i in range(2,12):
    data = ws.cell(row=i,column=2).value
    # print(data)
    
# Reading data from a range of cells
my_list = []

for value in ws.iter_rows(
    min_row=1, max_row=11, min_col=1, max_col=6,
    values_only=True):
    my_list.append(value)
    
for ele1,ele2,ele3,ele4,ele5,ele6 in my_list:
    # print(ele1,ele2,ele3,ele4,ele5,ele6)
    pass

# Writing to excel files
# Writing to a cell
ws['K1'] = 'Sum of Sales'

wb.save('videogamesales.xlsx')

# Appending new rows
new_row = (1,'The Legend of Zelda',1986,'Action','Nintendo',3.74,0.93,1.69,0.14,6.51,6.5)

ws.append(new_row)
    
wb.save('videogamesales.xlsx')


values = [ws.cell(row=ws.max_row,column=i).value for i in range(1,ws.max_column+1)]
# print(values)

# Deleting rows
ws.delete_rows(ws.max_row, 1)
wb.save('videogamesales.xlsx')

# Excel formulas
# Average
ws['P1'] = 'Average Sales'
ws['P2'] = '= AVERAGE(K2:K16220)'

# COUNTA 
ws['Q1'] = "Number of Populated Cells" 
ws['Q2'] = '=COUNTA(E2:E16220)'

# COUNTIF
ws['R1'] = 'Number of Rows with Sports Genre'
ws['R2'] = '=COUNTIF(E2:E16220, "Sports")'

# SUMIF
ws['S1'] = 'Total Sports Sales'
ws['S2'] = '=SUMIF(E2:E16220, "Sports",K2:K16220)'

# CEILING - rounds up to the nearest specified multiple
ws['T1'] = 'Rounded Sum of Sports Sales'
ws['T2'] = '=CEILING(S2,25)'

wb.save('videogamesales.xlsx')


# Addings charts to an excel file
ws = wb['Total Sales by Genre']

from openpyxl.chart import Reference

values = Reference(ws,           
                   min_col=2,  
                   max_col=2,  
                   min_row=1,  
                   max_row=13) 

cats = Reference(ws, 
                 min_col=1, 
                 max_col=1, 
                 min_row=2, 
                 max_row=13)

from openpyxl.chart import BarChart

chart = BarChart()
chart.add_data(values, titles_from_data=True)
chart.set_categories(cats)

# set the title of the chart
chart.title = "Total Sales"

# set the title of the x-axis
chart.x_axis.title = "Genre"

# set the title of the y-axis
chart.y_axis.title = "Total Sales by Genre"

ws.add_chart(chart,"D2")

# save the file 
wb.save("videogamesales.xlsx")





